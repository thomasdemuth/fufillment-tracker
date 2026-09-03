import io
import time

from openpyxl import load_workbook


def _refresh(c):
    r = c.post("/api/refresh", json={"all": True})
    jid = r.json()["job_id"]
    for _ in range(300):
        j = c.get(f"/api/jobs/{jid}").json()
        if j["status"] in ("done", "failed"):
            return j
        time.sleep(0.1)
    raise AssertionError("job stuck")


def test_attention_reasons(seeded):
    assert seeded.get("/api/attention").json() == []  # never polled -> nothing is stuck yet
    _refresh(seeded)
    rows = seeded.get("/api/attention").json()
    assert rows and all(r["reasons"] for r in rows)
    reasons = {x for r in rows for x in r["reasons"]}
    assert "exception" in reasons
    assert rows[0]["reasons"][0] in ("exception", "returned", "delivery_failed")
    only_exc = seeded.get("/api/attention", params={"status": "exception"}).json()
    assert all(r["status"] == "exception" for r in only_exc)


def test_export_csv_and_xlsx(seeded):
    r = seeded.get("/api/export", params={"format": "csv", "state": "CA"})
    assert r.status_code == 200 and r.headers["content-type"].startswith("text/csv")
    lines = r.text.strip().splitlines()
    total = seeded.get("/api/shipments", params={"state": "CA", "page_size": 1}).json()["total"]
    assert len(lines) == total + 1
    assert lines[0].startswith("Tracking number,Carrier,Status")
    r = seeded.get(
        "/api/export", params={"format": "xlsx", "columns": "tracking_number,recipient_name,status"}
    )
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Tracking number", "Recipient", "Status"]
    assert ws.max_row == 301
    assert seeded.get("/api/export", params={"format": "pdf"}).status_code == 422


def test_notes_and_tags(seeded):
    sid = seeded.get("/api/shipments", params={"page_size": 1}).json()["items"][0]["id"]
    d = seeded.post(f"/api/shipments/{sid}/notes", json={"body": "Called customer, will retry Monday"}).json()
    assert d["notes"][0]["body"].startswith("Called")
    nid = d["notes"][0]["id"]
    assert seeded.put(f"/api/notes/{nid}", json={"body": "edited"}).json()["body"] == "edited"
    d = seeded.put(f"/api/shipments/{sid}/tags", json={"tags": ["VIP", "gift", "VIP"]}).json()
    assert sorted(t["name"] for t in d["tags"]) == ["VIP", "gift"]
    tags = seeded.get("/api/tags").json()
    assert len(tags) == 2 and all(t["color"].startswith("#") for t in tags)
    vip = seeded.get("/api/shipments", params={"tag": "VIP"}).json()
    assert vip["total"] == 1 and vip["items"][0]["id"] == sid
    assert "VIP" in seeded.get("/api/export", params={"format": "csv", "tag": "VIP"}).text
    facets = seeded.get("/api/shipments/facets").json()
    assert [t["name"] for t in facets["tags"]] == ["VIP", "gift"]
    assert seeded.delete(f"/api/notes/{nid}").status_code == 204
    assert seeded.delete(f"/api/tags/{tags[0]['id']}").status_code == 204
    assert len(seeded.get(f"/api/shipments/{sid}").json()["tags"]) == 1
